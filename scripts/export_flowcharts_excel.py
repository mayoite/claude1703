from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def write_table(ws, headers: list[str], rows: list[tuple]):
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(14, max_len + 2), 60)


def main():
    root = Path(__file__).resolve().parents[1]
    out_path = root / "docs" / "ops" / "supabase-flowcharts.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    current_nodes = [
        ("S1", "AFC backups (manual corrections)", "Source Inputs"),
        ("S2", "seed_data.sql (product inserts)", "Source Inputs"),
        ("S3", "legacy image paths normalized into /images/catalog/*", "Source Inputs"),
        ("M1", "ad-hoc scripts + historical migrations", "Current Mapping Path"),
        ("M2", "products columns updated directly", "Current Mapping Path"),
        ("M3", "metadata patches applied per run", "Current Mapping Path"),
        ("T1", "products.category_id", "Current Storage"),
        ("T2", "products.series_id + series_name", "Current Storage"),
        ("T3", "products.metadata.subcategory/_slug", "Current Storage"),
        ("T4", "products.specs (jsonb)", "Current Storage"),
        ("T5", "products.flagship_image + images + scene_images", "Current Storage"),
        ("T6", "product_specs / product_images backfill tables", "Current Storage"),
        ("R1", "getCatalog/getProducts", "Current Serving"),
        ("R2", "optional Nhost fallback on Supabase error", "Current Serving"),
        ("R3", "filter API + grid rendering", "Current Serving"),
        (
            "E1",
            "duplicate slug/name variants; spec sparsity; image path/order drift; fallback masking DB errors",
            "Observed Issues",
        ),
    ]

    current_edges = [
        ("S1", "M1", ""),
        ("S2", "M1", ""),
        ("S3", "M1", ""),
        ("M1", "M2", ""),
        ("M2", "M3", ""),
        ("M3", "T1", ""),
        ("M3", "T2", ""),
        ("M3", "T3", ""),
        ("M3", "T4", ""),
        ("M3", "T5", ""),
        ("T4", "T6", ""),
        ("T5", "T6", ""),
        ("T1", "R1", ""),
        ("T2", "R1", ""),
        ("T3", "R1", ""),
        ("T4", "R1", ""),
        ("T5", "R1", ""),
        ("R1", "R2", ""),
        ("R2", "R3", ""),
        ("E1", "R3", "risk influence"),
    ]

    planned_nodes = [
        ("C1", "AFC backups JSON (authoritative corrections)", "Canonical Inputs"),
        ("C2", "seed_data.sql (fill missing specs/media)", "Canonical Inputs"),
        ("C3", "schema contracts (required fields)", "Canonical Inputs"),
        ("N1", "Parse all sources", "Normalize + Resolve"),
        ("N2", "Key by slug", "Normalize + Resolve"),
        ("N3", "Conflict on same field?", "Normalize + Resolve"),
        ("N4", "Resolve: AFC backup wins", "Normalize + Resolve"),
        ("N5", "Log diff by slug+field", "Normalize + Resolve"),
        ("F1", "Core: slug,name,description,category_id", "Field Lanes"),
        ("F2", "Series: series_id,series_name", "Field Lanes"),
        ("F3", "Metadata: subcategory,subcategory_slug,bifma,warranty", "Field Lanes"),
        ("F4", "Specs: dimensions,materials,features,sustainability_score", "Field Lanes"),
        ("F5", "Media: flagship_image,images[],scene_images[] + order", "Field Lanes"),
        ("K1", "Seating lane", "Category Branch"),
        ("K2", "Tables lane", "Category Branch"),
        ("K3", "Workstations lane", "Category Branch"),
        ("K4", "Soft-seating lane", "Category Branch"),
        ("K5", "Storages lane", "Category Branch"),
        ("K6", "Education lane", "Category Branch"),
        ("Q1", "Required fields complete?", "Validation Gates"),
        ("Q2", "Unique slug + category mapping valid?", "Validation Gates"),
        ("Q3", "Specs coverage threshold met?", "Validation Gates"),
        ("Q4", "Image URL reachable + ordered?", "Validation Gates"),
        ("Q5", "Category branch constraints valid?", "Validation Gates"),
        ("D1", "Build staging dataset", "Promotion"),
        ("D2", "Run full product/page audit", "Promotion"),
        ("D3", "Audit clean?", "Promotion"),
        ("D4", "Transactional production upsert", "Promotion"),
        ("D5", "Publish report + lock mapping snapshot", "Promotion"),
        ("D6", "Rollback to backup snapshot", "Promotion"),
    ]

    planned_edges = [
        ("C1", "N1", ""),
        ("C2", "N1", ""),
        ("C3", "N1", ""),
        ("N1", "N2", ""),
        ("N2", "N3", ""),
        ("N3", "F1", "No conflict"),
        ("N3", "N4", "Conflict"),
        ("N4", "N5", ""),
        ("N5", "F1", ""),
        ("F1", "F2", ""),
        ("F2", "F3", ""),
        ("F3", "F4", ""),
        ("F4", "F5", ""),
        ("F5", "K1", ""),
        ("F5", "K2", ""),
        ("F5", "K3", ""),
        ("F5", "K4", ""),
        ("F5", "K5", ""),
        ("F5", "K6", ""),
        ("K1", "Q1", ""),
        ("K2", "Q1", ""),
        ("K3", "Q1", ""),
        ("K4", "Q1", ""),
        ("K5", "Q1", ""),
        ("K6", "Q1", ""),
        ("Q1", "Q2", ""),
        ("Q2", "Q3", ""),
        ("Q3", "Q4", ""),
        ("Q4", "Q5", ""),
        ("Q5", "D1", ""),
        ("D1", "D2", ""),
        ("D2", "D3", ""),
        ("D3", "D4", "Yes"),
        ("D4", "D5", ""),
        ("D3", "D6", "No"),
    ]

    field_mapping = [
        (
            "category_id",
            "seed_data.sql + ad-hoc backup patches",
            "products.category_id",
            "Canonical from AFC backups, seed SQL fills gaps",
            "Unique slug/category mapping valid",
        ),
        (
            "series_id/series_name",
            "seed_data.sql",
            "products.series_id, products.series_name",
            "Normalize by slug and category branch rules",
            "Required fields complete",
        ),
        (
            "metadata.subcategory/subcategory_slug",
            "AFC backups + seed_data.sql",
            "products.metadata",
            "AFC backup priority; explicit conflict log",
            "Category branch constraints valid",
        ),
        (
            "specs.dimensions/materials/features",
            "seed_data.sql (primary) + AFC checks",
            "products.specs + product_specs",
            "Populate missing specs before promotion",
            "Specs coverage threshold met",
        ),
        (
            "media paths/order",
            "legacy catalog image paths + seed_data.sql",
            "products.flagship_image/images/scene_images + product_images",
            "Normalize path + order; keep deterministic gallery sort",
            "Image URL reachable + ordered",
        ),
        (
            "fallback behavior",
            "runtime logic in lib/getProducts.ts",
            "getCatalog/getProducts response path",
            "Do not treat fallback as canonical DB truth",
            "Post-cutover product/page audit",
        ),
    ]

    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    overview_rows = [
        ("Workbook", "Supabase Flowchart Conversion"),
        ("Generated At", datetime.now().isoformat(timespec="seconds")),
        ("Current Chart HTML", "docs/ops/current-mapping-chart.html"),
        ("Planned Chart HTML", "docs/ops/planned-mapping-chart.html"),
        ("Current Nodes", len(current_nodes)),
        ("Current Edges", len(current_edges)),
        ("Planned Nodes", len(planned_nodes)),
        ("Planned Edges", len(planned_edges)),
    ]
    write_table(overview, ["Key", "Value"], overview_rows)

    ws_current_nodes = wb.create_sheet("Current_Nodes")
    write_table(ws_current_nodes, ["Node ID", "Label", "Group"], current_nodes)

    ws_current_edges = wb.create_sheet("Current_Edges")
    write_table(ws_current_edges, ["From", "To", "Condition/Note"], current_edges)

    ws_planned_nodes = wb.create_sheet("Planned_Nodes")
    write_table(ws_planned_nodes, ["Node ID", "Label", "Group"], planned_nodes)

    ws_planned_edges = wb.create_sheet("Planned_Edges")
    write_table(ws_planned_edges, ["From", "To", "Condition/Note"], planned_edges)

    ws_field_mapping = wb.create_sheet("Field_Mapping")
    write_table(
        ws_field_mapping,
        ["Field Group", "Current Source", "Current Store", "Planned Rule", "Validation Gate"],
        field_mapping,
    )

    wb.save(out_path)
    print(str(out_path))


if __name__ == "__main__":
    main()
